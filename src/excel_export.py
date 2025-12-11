import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from database.models import Assignment, Employee, Object, CashWithdrawal, db
from datetime import datetime
import os

def export_to_excel(data, filename):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Табель"
    
    # Стили
    header_font = Font(name="Times New Roman", size=12, bold=True, color="FFFFFF")
    data_font = Font(name="Times New Roman", size=12, color="000000")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    orange_fill = PatternFill(start_color="FFE4B5", end_color="FFE4B5", fill_type="solid")  # светло-оранжевый
    green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")   # светло-зеленый
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # белый
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                   top=Side(style='thin'), bottom=Side(style='thin'))
    center_align = Alignment(horizontal='center', vertical='center')
    
    # Название объекта (до последнего числа месяца)
    ws.merge_cells('A1:AO1')
    ws['A1'] = data['title']
    ws['A1'].font = Font(name="Times New Roman", size=14, bold=True)
    ws['A1'].alignment = center_align
    
    # Адрес объекта
    ws.merge_cells('A2:AO2')
    ws['A2'] = data.get('address', '')
    ws['A2'].font = Font(name="Times New Roman", size=12)
    ws['A2'].alignment = center_align
    
    # Заголовки колонок (без пробела)
    row = 3
    ws['A3'] = "Служба охраны"
    ws['B3'] = "РАЗРЯД"
    
    # Дни месяца
    for i, day in enumerate(range(1, 32), 3):
        col = get_column_letter(i)
        ws[f'{col}{row}'] = day
    
    # Остальные заголовки
    headers = ["Итого часов", "Итого в гроссе", "Тарифная ставка", 
               "Поощрение", "На карту", "ВЗН", "Удержания", "Итого на руки"]
    for i, header in enumerate(headers, 34):
        col = get_column_letter(i)
        ws[f'{col}{row}'] = header
    
    # Применяем стили к заголовкам
    for col in range(1, 42):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_align
    
    # Данные сотрудников
    for idx, employee in enumerate(data['rows'], 1):
        row_num = row + idx
        ws[f'A{row_num}'] = employee['name']
        # Записываем разряд из разных возможных ключей
        guard_rank = employee.get('разряд') or employee.get('guard_rank') or employee.get('Разряд') or ''
        ws[f'B{row_num}'] = guard_rank
        
        # Дни
        days_data = employee.get('days', {})
        for day in range(1, 32):
            col = get_column_letter(day + 2)
            ws[f'{col}{row_num}'] = days_data.get(str(day), '')
        
        # Итоговые данные
        ws[f'AH{row_num}'] = employee.get('итого_часов', 0)
        ws[f'AI{row_num}'] = employee.get('итого_в_гроссе', 0)
        ws[f'AJ{row_num}'] = employee.get('тарифная_ставка', 0)
        ws[f'AK{row_num}'] = employee.get('поощрение', 0)
        ws[f'AL{row_num}'] = employee.get('на_карту', 0)
        ws[f'AM{row_num}'] = employee.get('взн', 0)
        ws[f'AN{row_num}'] = employee.get('удержания', 0)
        ws[f'AO{row_num}'] = employee.get('итого_на_руки', 0)
        
        # Границы, шрифт и цвета для строк данных
        for col in range(1, 42):
            cell = ws.cell(row=row_num, column=col)
            cell.border = border
            cell.alignment = center_align
            cell.font = data_font
            
            # Цветовая схема по колонкам
            if col == 1:  # Служба охраны - светло-оранжевый
                cell.fill = orange_fill
            elif col == 2:  # Разряд - светло-зеленый
                cell.fill = green_fill
            else:  # Остальные колонки белые
                cell.fill = white_fill
    
    # Итоговая строка с суммой часов по дням
    footer_row = row + len(data['rows']) + 1
    ws[f'A{footer_row}'] = "ИТОГО:"
    ws[f'A{footer_row}'].font = Font(name="Times New Roman", size=12, bold=True)
    
    # Подсчет итогов по дням
    for day in range(1, 32):
        col = get_column_letter(day + 2)
        day_total = 0
        for employee in data['rows']:
            days_data = employee.get('days', {})
            day_total += days_data.get(str(day), 0)
        ws[f'{col}{footer_row}'] = day_total if day_total > 0 else ''
    
    footer = data.get('footer', {})
    ws[f'AH{footer_row}'] = footer.get('итого_часов', 0)
    ws[f'AI{footer_row}'] = footer.get('итого_в_гроссе', 0)
    ws[f'AK{footer_row}'] = footer.get('поощрение', 0)
    ws[f'AL{footer_row}'] = footer.get('на_карту', 0)
    ws[f'AM{footer_row}'] = footer.get('взн', 0)
    ws[f'AN{footer_row}'] = footer.get('удержания', 0)
    ws[f'AO{footer_row}'] = footer.get('итого_на_руки', 0)
    
    # Стили для итоговой строки
    for col in range(1, 42):
        cell = ws.cell(row=footer_row, column=col)
        cell.border = border
        cell.font = Font(name="Times New Roman", size=12, bold=True)
        cell.alignment = center_align
        
        # Цвета для итоговой строки
        if col == 1:  # Служба охраны - оранжевый
            cell.fill = orange_fill
        elif col == 2:  # Разряд - зеленый
            cell.fill = green_fill
        else:
            cell.fill = white_fill
    
    # Настройка размеров колонок
    ws.column_dimensions['A'].width = 40  # ФИО
    ws.column_dimensions['B'].width = 12  # Разряд
    
    # Дни месяца (узкие колонки)
    for day in range(1, 32):
        col = get_column_letter(day + 2)
        ws.column_dimensions[col].width = 4
    
    # Итоговые колонки
    ws.column_dimensions['AH'].width = 15  # итого часов
    ws.column_dimensions['AI'].width = 16  # итого в гроссе
    ws.column_dimensions['AJ'].width = 20  # тарифная ставка
    ws.column_dimensions['AK'].width = 14  # поощрение
    ws.column_dimensions['AL'].width = 12  # на карту
    ws.column_dimensions['AM'].width = 10  # ВЗН
    ws.column_dimensions['AN'].width = 13  # удержания
    ws.column_dimensions['AO'].width = 16  # итого на руки
    
    wb.save(filename)

def export_assignments_to_excel(month=None, year=None):
    """Экспортирует назначения за указанный месяц в Excel файлы по объектам"""
    try:
        if db.is_closed():
            db.connect()
        
        # Если месяц и год не указаны, используем текущие
        if month is None or year is None:
            current_date = datetime.now()
            month = current_date.month
            year = current_date.year
        
        # Фильтруем назначения по месяцу и году
        assignments = Assignment.select(Assignment, Employee, Object).join(Employee).switch(Assignment).join(Object).where(
            (Assignment.date.year == year) & (Assignment.date.month == month)
        )
        
        # Получаем ВЗН за тот же период
        cash_withdrawals = CashWithdrawal.select(CashWithdrawal, Employee, Object).join(Employee).switch(CashWithdrawal).join(Object).where(
            (CashWithdrawal.date.year == year) & (CashWithdrawal.date.month == month)
        )
        
        if not assignments.exists():
            return False, "Нет данных для экспорта. Добавьте сотрудников и назначения."
        
        objects_data = {}
        
        # Обрабатываем обычные назначения
        for assignment in assignments:
            obj_name = assignment.object.name
            if obj_name not in objects_data:
                objects_data[obj_name] = []
            
            day = assignment.date.day
            employee_found = False
            for emp_data in objects_data[obj_name]:
                if emp_data['name'] == assignment.employee.full_name:
                    emp_data['разряд'] = assignment.employee.guard_rank
                    emp_data['days'][str(day)] = assignment.hours
                    emp_data['итого_часов'] += assignment.hours
                    emp_data['поощрение'] += float(assignment.bonus_amount)
                    emp_data['удержания'] += float(assignment.deduction_amount)
                    emp_data['итого_в_гроссе'] = emp_data['итого_часов'] * emp_data['тарифная_ставка']
                    # Инициализируем ВЗН, если его нет
                    if 'взн' not in emp_data:
                        emp_data['взн'] = 0
                    payment_method = getattr(assignment.employee, 'payment_method', 'на карту')
                    # ВЗН не входит в итоговую сумму выдачи
                    final_salary = emp_data['итого_в_гроссе'] - emp_data['удержания'] + emp_data['поощрение']
                    emp_data['на_карту'] = final_salary if payment_method == 'на карту' else 0
                    emp_data['итого_на_руки'] = final_salary if payment_method == 'на руки' else 0
                    employee_found = True
                    break
            
            if not employee_found:
                total_salary = assignment.hours * float(assignment.hourly_rate)
                payment_method = getattr(assignment.employee, 'payment_method', 'на карту')
                bonus = float(assignment.bonus_amount)
                deductions = float(assignment.deduction_amount)
                # ВЗН не входит в итоговую сумму выдачи
                final_salary = total_salary - deductions + bonus
                
                objects_data[obj_name].append({
                    'name': assignment.employee.full_name,
                    'разряд': assignment.employee.guard_rank,
                    'days': {str(day): assignment.hours},
                    'итого_часов': assignment.hours,
                    'итого_в_гроссе': total_salary,
                    'тарифная_ставка': float(assignment.hourly_rate),
                    'поощрение': bonus,
                    'удержания': deductions,
                    'взн': 0,
                    'на_карту': final_salary if payment_method == 'на карту' else 0,
                    'итого_на_руки': final_salary if payment_method == 'на руки' else 0
                })
        
        # Обрабатываем ВЗН отдельно
        for cash_withdrawal in cash_withdrawals:
            obj_name = cash_withdrawal.object.name
            if obj_name not in objects_data:
                objects_data[obj_name] = []
            
            employee_found = False
            for emp_data in objects_data[obj_name]:
                if emp_data['name'] == cash_withdrawal.employee.full_name:
                    # Добавляем ВЗН к существующему сотруднику
                    vzn_amount = cash_withdrawal.hours * float(cash_withdrawal.hourly_rate)
                    emp_data['взн'] += vzn_amount
                    emp_data['поощрение'] += float(cash_withdrawal.bonus_amount)
                    emp_data['удержания'] += float(cash_withdrawal.deduction_amount)
                    # Добавляем часы ВЗН в соответствующий день
                    day = str(cash_withdrawal.date.day)
                    if day in emp_data['days']:
                        emp_data['days'][day] += cash_withdrawal.hours
                    else:
                        emp_data['days'][day] = cash_withdrawal.hours
                    emp_data['итого_часов'] += cash_withdrawal.hours
                    employee_found = True
                    break
            
            if not employee_found:
                # Создаем новую запись только для ВЗН
                vzn_amount = cash_withdrawal.hours * float(cash_withdrawal.hourly_rate)
                payment_method = getattr(cash_withdrawal.employee, 'payment_method', 'на карту')
                day = cash_withdrawal.date.day
                
                objects_data[obj_name].append({
                    'name': cash_withdrawal.employee.full_name,
                    'разряд': cash_withdrawal.employee.guard_rank,
                    'days': {str(day): cash_withdrawal.hours},
                    'итого_часов': cash_withdrawal.hours,
                    'итого_в_гроссе': 0,
                    'тарифная_ставка': float(cash_withdrawal.hourly_rate),
                    'поощрение': float(cash_withdrawal.bonus_amount),
                    'удержания': float(cash_withdrawal.deduction_amount),
                    'взн': vzn_amount,
                    'на_карту': 0,
                    'итого_на_руки': 0
                })
        
        month_names = {
            1: 'Январь', 2: 'Февраль', 3: 'Март', 4: 'Апрель',
            5: 'Май', 6: 'Июнь', 7: 'Июль', 8: 'Август',
            9: 'Сентябрь', 10: 'Октябрь', 11: 'Ноябрь', 12: 'Декабрь'
        }
        month_name = month_names[month]
        
        for obj_name, rows in objects_data.items():
            try:
                obj_record = Object.get(Object.name == obj_name)
                obj_address = obj_record.address or ''
            except:
                obj_address = ''
            
            data = {
                'title': obj_name,
                'address': obj_address,
                'rows': rows,
                'footer': {
                    'итого_часов': sum(r['итого_часов'] for r in rows),
                    'итого_в_гроссе': sum(r['итого_в_гроссе'] for r in rows),
                    'поощрение': sum(r['поощрение'] for r in rows),
                    'на_карту': sum(r['на_карту'] for r in rows),
                    'взн': sum(r['взн'] for r in rows),
                    'удержания': sum(r['удержания'] for r in rows),
                    'итого_на_руки': sum(r['итого_на_руки'] for r in rows)
                }
            }
            
            filename = f"{obj_name} {month_name} {year}.xlsx"
            filepath = os.path.join(os.path.expanduser('~'), 'Downloads', filename)
            export_to_excel(data, filepath)
        
        return True, f"Данные экспортированы в Загрузки"
        
    except Exception as ex:
        return False, f"Ошибка экспорта: {str(ex)}"
    finally:
        if not db.is_closed():
            db.close()
